# import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Only work with extension xlsx
# Open the Workbook. If same directory you do not need to specify the path.
# If different folder you need to add the path. C:/Users/msosa/Desktop/Python/Test/Grades.xlsx
# Check how to map the folder.
wb = load_workbook('Grades.xlsx')

# Open the worksheet
ws = wb.active
print(ws)

# Accessing cell values.
print(ws['A1'].value)

# To change a cell value. And you need to save the workbook - wb
ws['A2'].value = "test"

# To save wb.save('name.xlsx'), if file is open you will get an error.
wb.save('Grades.xlsx')

# To access different sheets, that are no the active.
# To see all the sheets.
print(wb.sheetnames)

# To access another sheet (no active) print(wb['name'])
# print(wb['Sheet1'])

# Create sheet
wb.create_sheet("Test")
print(wb.sheetnames)

# To create a new workbook !!!
wb = Workbook()
ws = wb.active
ws.title = "Data"

# The following will take a lot of time to add values.
ws['A1'] =
ws['B1'] =

# To add one row you can use:
ws.append(['Tim', 'Is', 'great', '!'])

wb.save(filename)

# ************ To loop **************

for row in range(1, 100):
    for col in range(1, 5):
        # char = chr(65 + col)
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

# Insert empty row and columns
ws.insert_rows(7)
ws.insert_cols(2)

# Delete rows and columns
ws.delete_rows(7)
ws.delete_cols(2)

# to move rows and columns
ws.move_range("C1:D11", rows=2, cols=2)

# To insert values

# This is a list that you get from the data, or you can create the list manually.
headings = ['Name'] + list(data['Joe'].keys())

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)


# max_row / max_column to count the number of rows / columns.
wb = load_workbook(path, use_iterators=True)
sheet = wb.worksheets[0]

row_count = sheet.max_row
column_count = sheet.max_column

import openpyxl as xl

wb = xl.load_workbook("Sample.xlsx", enumerate)

#the 2 lines under do the same.
sheet = wb.get_sheet_by_name('sheet')
sheet = wb.worksheets[0]

row_count = sheet.max_row
column_count = sheet.max_column


# f string allow to string different things together inside the string.
#print(f'{ws["A2"].value}:{ws["B2"].value}')

#name = ws["A2"].value
#color = ws["B2"].value
#print(f'{name}: {color}')